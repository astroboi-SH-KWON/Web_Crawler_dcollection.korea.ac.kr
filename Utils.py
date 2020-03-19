import openpyxl
from time import clock

class Utils:
    def __init__(self):
        self.ext_excel = ".xlsx"

    def make_excel(self, data):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        col = 1
        header_list = []
        for header in data[1][0].keys():
            header_list.append(header)
            sheet.cell(row=row, column=col, value=header)
            col = col + 1

        row = 2
        for tmp_dict in data[1]:
            tmp_col = 1
            for header in header_list:
                sheet.cell(row=row, column=tmp_col, value=tmp_dict[header])
                tmp_col = tmp_col + 1
            row = row + 1

        workbook.save(filename=data[0] + str(clock()) + self.ext_excel)